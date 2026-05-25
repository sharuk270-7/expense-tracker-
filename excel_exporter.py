"""
Excel export functionality for expense data
Exports expenses to .xlsx format with formatting, summaries, and charts
"""
import os
import re
import sqlite3
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import ExpenseDatabase
from config import CURRENCY, EXPENSE_PATTERNS

class ExcelExporter:
    def __init__(self):
        self.db = ExpenseDatabase()
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # timezones
        self._utc = ZoneInfo("UTC")
        self._ist = ZoneInfo("Asia/Kolkata")

    def _extract_pattern_list(self, description, category=None):
        """Extract matched pattern keywords from description/category text."""
        text = f"{description or ''} {category or ''}".lower()
        # Remove receipt metadata fragments so only expense-pattern terms remain.
        text = re.sub(r"\bqty\s*:\s*[0-9]+(?:\.[0-9]+)?\b", " ", text, flags=re.IGNORECASE)
        text = re.sub(r"\bunit\s*:\s*[0-9]+(?:\.[0-9]+)?\b", " ", text, flags=re.IGNORECASE)
        text = re.sub(r"\b(qty|quantity|unit)\b", " ", text, flags=re.IGNORECASE)
        text = re.sub(r"\s{2,}", " ", text).strip()
        if not text.strip():
            return []

        matches = []
        for _, keywords in EXPENSE_PATTERNS.items():
            for keyword in keywords:
                kw = (keyword or "").strip().lower()
                if not kw:
                    continue
                if re.search(rf"\b{re.escape(kw)}\b", text):
                    if kw not in matches:
                        matches.append(kw)
        return matches

    def _extract_pattern_names(self, description, category=None):
        patterns = self._extract_pattern_list(description, category)
        return ", ".join(patterns) if patterns else "Other"

    def _extract_image_item_name(self, description):
        """Return clean item name from image receipt description."""
        name = (description or "").strip()
        if not name:
            return "Receipt Item"

        # Legacy rows may have "Item | Qty: x | Unit: y"
        name = name.split("|")[0].strip()
        name = re.sub(r"\bqty\s*:\s*[0-9]+(?:\.[0-9]+)?\b", " ", name, flags=re.IGNORECASE)
        name = re.sub(r"\bquantity\s*:\s*[0-9]+(?:\.[0-9]+)?\b", " ", name, flags=re.IGNORECASE)
        name = re.sub(r"\bunit\s*:\s*[0-9]+(?:\.[0-9]+)?\b", " ", name, flags=re.IGNORECASE)
        name = re.sub(r"\s{2,}", " ", name).strip(" |-")
        return name or "Receipt Item"

    def _description_for_excel(self, description, category=None, source=None):
        """Description output strategy for Excel sheets."""
        if (source or "").strip().lower() == "image":
            return self._extract_image_item_name(description)
        return self._extract_pattern_names(description, category)

    def _parse_to_ist(self, date_str):
        """Parse ISO datetime string (assumed UTC if naive) and convert to IST."""
        date_obj = datetime.fromisoformat(date_str)
        if date_obj.tzinfo is None:
            date_obj = date_obj.replace(tzinfo=self._utc)
        return date_obj.astimezone(self._ist)

    def _normalize_txn_datetime_display(self, txn_time, logged_date_obj):
        """
        Ensure UPI transaction datetime is always displayed as date + time.
        If txn_time has only date/time/empty, complete missing parts from logged timestamp.
        """
        logged_display = logged_date_obj.strftime("%d-%m-%Y %H:%M")
        raw = (txn_time or "").strip()
        if not raw:
            return logged_display

        value = re.sub(r"\s+", " ", raw)
        has_date = bool(
            re.search(
                r"\b\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4}\b|\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b",
                value,
                flags=re.IGNORECASE,
            )
        )
        has_time = bool(
            re.search(
                r"\b\d{1,2}:\d{2}(?:\s*(?:am|pm))?\b",
                value,
                flags=re.IGNORECASE,
            )
        )

        if has_date and has_time:
            return value
        if has_date and not has_time:
            return f"{value} {logged_date_obj.strftime('%H:%M')}"
        if has_time and not has_date:
            return f"{logged_date_obj.strftime('%d-%m-%Y')} {value}"
        return logged_display

    def _get_upi_meta_map(self, user_id, expense_ids):
        """Fetch UPI metadata for given expense IDs."""
        if not expense_ids:
            return {}

        conn = sqlite3.connect(self.db.db_path)
        cursor = conn.cursor()
        placeholders = ",".join("?" for _ in expense_ids)
        query = f"""
            SELECT id, source, payment_method, amount, upi_to, upi_from, transaction_time, transaction_id
            FROM expenses
            WHERE user_id = ?
              AND id IN ({placeholders})
        """
        cursor.execute(query, (user_id, *expense_ids))
        rows = cursor.fetchall()
        conn.close()

        meta_map = {}
        for row in rows:
            exp_id, source, payment_method, amount, upi_to, upi_from, transaction_time, transaction_id = row
            meta_map[exp_id] = {
                "source": source,
                "payment_method": payment_method,
                "amount": amount,
                "upi_to": upi_to,
                "upi_from": upi_from,
                "transaction_time": transaction_time,
                "transaction_id": transaction_id,
            }
        return meta_map

    def _extract_upi_excel_columns(self, meta, fallback_amount, logged_date_obj):
        """Return tuple for columns G-K: To, From, Amount, Date/Time, UPI transaction id."""
        if not meta:
            return "", "", "", "", ""

        source = (meta.get("source") or "").strip().lower()
        payment_method = (meta.get("payment_method") or "").strip().lower()
        is_upi = (
            source == "online_payment"
            or payment_method == "upi"
            or bool(meta.get("upi_to"))
            or bool(meta.get("upi_from"))
            or bool(meta.get("transaction_time"))
        )
        if not is_upi:
            return "", "", "", "", ""

        upi_amount = meta.get("amount")
        if upi_amount is None:
            upi_amount = fallback_amount

        txn_display = self._normalize_txn_datetime_display(
            meta.get("transaction_time"),
            logged_date_obj,
        )

        return (
            meta.get("upi_to") or "",
            meta.get("upi_from") or "",
            upi_amount if upi_amount is not None else "",
            txn_display,
            meta.get("transaction_id") or "",
        )
    
    def export_all_expenses(self, user_id, filename=None):
        """Export all user expenses to Excel"""
        if not filename:
            filename = f"expenses_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "All Expenses"
        
        # Get all expenses
        expenses = self.db.get_expenses(user_id)
        
        if not expenses:
            ws['A1'] = "No expenses found"
            wb.save(filename)
            return filename
        
        # Headers
        headers = [
            "ID",
            "Date",
            "Category",
            "Amount",
            "Description",
            "Source",
            "To",
            "From",
            "Amount",
            "Date/Time",
            "UPI transaction id",
        ]
        self._add_headers(ws, headers)

        upi_meta_map = self._get_upi_meta_map(user_id, [exp_id for exp_id, *_ in expenses])

        # Data
        # Use sequential Excel IDs (1..N) so numbering restarts after deletions
        for seq, (exp_id, amount, category, description, date, source) in enumerate(expenses, start=1):
            row_idx = seq + 1
            date_obj = self._parse_to_ist(date)
            to_value, from_value, upi_amount, upi_date_time, upi_txn_id = self._extract_upi_excel_columns(
                upi_meta_map.get(exp_id, {}),
                amount,
                date_obj,
            )
            ws[f'A{row_idx}'] = seq
            ws[f'B{row_idx}'] = date_obj.strftime("%d-%m-%Y %H:%M")
            ws[f'C{row_idx}'] = category
            ws[f'D{row_idx}'] = amount
            ws[f'E{row_idx}'] = self._description_for_excel(description, category, source)
            ws[f'F{row_idx}'] = source if source else "text"
            ws[f'G{row_idx}'] = to_value
            ws[f'H{row_idx}'] = from_value
            ws[f'I{row_idx}'] = upi_amount
            ws[f'J{row_idx}'] = upi_date_time
            ws[f'K{row_idx}'] = upi_txn_id

            # Format currency column
            ws[f'D{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
            if upi_amount != "":
                ws[f'I{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'

            # Apply border
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                ws[f'{col}{row_idx}'].border = self.thin_border
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 46
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 50
        ws.column_dimensions['H'].width = 50
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 26
        ws.column_dimensions['K'].width = 50
        
        # Add summary sheet
        self._add_summary_sheet(wb, user_id, expenses)
        
        # Add monthly breakdown sheet
        self._add_monthly_breakdown(wb, user_id, expenses)

        # Add bill totals sheet (subtotal/total/grand total entries from receipts)
        self._add_bill_totals_sheet(wb, user_id, days=None, sheet_name="Bill Totals")

        # Add bill analysis sheet (category + subtotal/total/grand total/amount)
        self._add_bill_analysis_sheet(wb, user_id, days=None, sheet_name="Bill Analysis")

        # Add bill items sheet (item-level category/amount/quantity)
        self._add_bill_items_sheet(wb, user_id, days=None, sheet_name="Bill Items")

        # Add UPI details sheet (to/from/amount/date-time/transaction ID)
        self._add_upi_details_sheet(wb, user_id, days=None, sheet_name="UPI Details")

        # Add pattern summary sheet
        self._add_pattern_summary_sheet(wb, expenses, sheet_name="Pattern Summary")
        
        wb.save(filename)
        return filename
    
    def export_monthly_expenses(self, user_id, filename=None):
        """Export expenses for the current month"""
        if not filename:
            filename = f"expenses_monthly_{user_id}_{datetime.now().strftime('%Y%m_%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly Expenses"
        
        # Get monthly expenses
        expenses = self.db.get_summary(user_id, days=30)
        all_expenses = self.db.get_expenses(user_id, days=30)
        
        if not all_expenses:
            ws['A1'] = "No expenses found for this month"
            wb.save(filename)
            return filename
        
        # Headers
        headers = ["Category", "Total Amount", "Transaction Count", "Average Per Item"]
        self._add_headers(ws, headers)
        
        # Data
        for row_idx, (category, total, count) in enumerate(expenses, start=2):
            avg = total / count if count > 0 else 0
            ws[f'A{row_idx}'] = category
            ws[f'B{row_idx}'] = total
            ws[f'C{row_idx}'] = count
            ws[f'D{row_idx}'] = avg
            
            # Format currency columns
            ws[f'B{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
            ws[f'D{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
            
            # Apply border
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_idx}'].border = self.thin_border
        
        # Add total row
        total_row = len(expenses) + 2
        ws[f'A{total_row}'] = "TOTAL"
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'] = f"=SUM(B2:B{total_row - 1})"
        ws[f'B{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'].number_format = f'"{CURRENCY}"#,##0.00'
        ws[f'B{total_row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        for col in ['A', 'B']:
            ws[f'{col}{total_row}'].border = self.thin_border
        
        # Column widths
        ws.column_dimensions['A'].width = 26
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 24
        
        # Add detailed transactions sheet
        self._add_detailed_sheet(wb, all_expenses, user_id=user_id)

        # Add bill totals sheet for this period
        self._add_bill_totals_sheet(wb, user_id, days=30, sheet_name="Bill Totals")

        # Add bill analysis sheet for this period
        self._add_bill_analysis_sheet(wb, user_id, days=30, sheet_name="Bill Analysis")

        # Add bill items sheet for this period
        self._add_bill_items_sheet(wb, user_id, days=30, sheet_name="Bill Items")

        # Add UPI details sheet for this period
        self._add_upi_details_sheet(wb, user_id, days=30, sheet_name="UPI Details")

        # Add pattern summary sheet for this period
        self._add_pattern_summary_sheet(wb, all_expenses, sheet_name="Pattern Summary")
        
        wb.save(filename)
        return filename
    
    def export_custom_period(self, user_id, days, filename=None):
        """Export expenses for a custom period"""
        if not filename:
            filename = f"expenses_{days}days_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Last {days} Days"
        
        # Get expenses for period
        all_expenses = self.db.get_expenses(user_id, days=days)
        summary = self.db.get_summary(user_id, days=days)
        
        if not all_expenses:
            ws['A1'] = f"No expenses found in the last {days} days"
            wb.save(filename)
            return filename
        
        # Summary section
        ws['A1'] = f"Expense Summary - Last {days} Days"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        
        # Summary headers
        ws['A3'] = "Category"
        ws['B3'] = "Total"
        ws['C3'] = "Count"
        ws['D3'] = "Avg/Item"
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}3'].font = Font(bold=True)
            ws[f'{col}3'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws[f'{col}3'].border = self.thin_border
        
        # Summary data
        total_sum = 0
        for row_idx, (category, total, count) in enumerate(summary, start=4):
            avg = total / count if count > 0 else 0
            ws[f'A{row_idx}'] = category
            ws[f'B{row_idx}'] = total
            ws[f'C{row_idx}'] = count
            ws[f'D{row_idx}'] = avg
            total_sum += total
            
            ws[f'B{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
            ws[f'D{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_idx}'].border = self.thin_border
        
        # Total row
        total_row = len(summary) + 4
        ws[f'A{total_row}'] = "TOTAL"
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'] = total_sum
        ws[f'B{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'].number_format = f'"{CURRENCY}"#,##0.00'
        ws[f'B{total_row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        
        for col in ['A', 'B']:
            ws[f'{col}{total_row}'].border = self.thin_border
        
        # Column widths
        ws.column_dimensions['A'].width = 26
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 24
        
        # Add detailed transactions sheet
        self._add_detailed_sheet(wb, all_expenses, sheet_name=f"Details - {days}d", user_id=user_id)

        # Add bill totals sheet for this period
        self._add_bill_totals_sheet(wb, user_id, days=days, sheet_name=f"Bill Totals {days}d")

        # Add bill analysis sheet for this period
        self._add_bill_analysis_sheet(wb, user_id, days=days, sheet_name=f"Bill Analysis {days}d")

        # Add bill items sheet for this period
        self._add_bill_items_sheet(wb, user_id, days=days, sheet_name=f"Bill Items {days}d")

        # Add UPI details sheet for this period
        self._add_upi_details_sheet(wb, user_id, days=days, sheet_name=f"UPI Details {days}d")

        # Add pattern summary sheet for this period
        self._add_pattern_summary_sheet(wb, all_expenses, sheet_name=f"Pattern Summary {days}d")
        
        wb.save(filename)
        return filename

    def export_date_range(self, user_id, start_date, end_date, filename=None):
        """Export expenses for a custom inclusive date range (YYYY-MM-DD to YYYY-MM-DD)."""
        if not filename:
            start_token = start_date.replace("-", "")
            end_token = end_date.replace("-", "")
            filename = f"expenses_range_{start_token}_{end_token}_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Custom Range"

        all_expenses = self.db.get_expenses_date_range(user_id, start_date, end_date)
        summary = self.db.get_summary_date_range(user_id, start_date, end_date)

        if not all_expenses:
            ws['A1'] = f"No expenses found between {start_date} and {end_date}"
            wb.save(filename)
            return filename

        ws['A1'] = f"Expense Summary - {start_date} to {end_date}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")

        ws['A3'] = "Category"
        ws['B3'] = "Total"
        ws['C3'] = "Count"
        ws['D3'] = "Avg/Item"

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}3'].font = Font(bold=True)
            ws[f'{col}3'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws[f'{col}3'].border = self.thin_border

        total_sum = 0
        for row_idx, (category, total, count) in enumerate(summary, start=4):
            avg = total / count if count > 0 else 0
            ws[f'A{row_idx}'] = category
            ws[f'B{row_idx}'] = total
            ws[f'C{row_idx}'] = count
            ws[f'D{row_idx}'] = avg
            total_sum += total

            ws[f'B{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
            ws[f'D{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'

            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_idx}'].border = self.thin_border

        total_row = len(summary) + 4
        ws[f'A{total_row}'] = "TOTAL"
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'] = total_sum
        ws[f'B{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'].number_format = f'"{CURRENCY}"#,##0.00'
        ws[f'B{total_row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        for col in ['A', 'B']:
            ws[f'{col}{total_row}'].border = self.thin_border

        ws.column_dimensions['A'].width = 26
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 24

        self._add_detailed_sheet(wb, all_expenses, sheet_name="Details - Range", user_id=user_id)
        self._add_bill_totals_sheet(
            wb,
            user_id,
            start_date=start_date,
            end_date=end_date,
            sheet_name="Bill Totals Range",
        )
        self._add_bill_analysis_sheet(
            wb,
            user_id,
            start_date=start_date,
            end_date=end_date,
            sheet_name="Bill Analysis Range",
        )
        self._add_bill_items_sheet(
            wb,
            user_id,
            start_date=start_date,
            end_date=end_date,
            sheet_name="Bill Items Range",
        )
        self._add_upi_details_sheet(
            wb,
            user_id,
            start_date=start_date,
            end_date=end_date,
            sheet_name="UPI Details Range",
        )
        self._add_pattern_summary_sheet(wb, all_expenses, sheet_name="Pattern Summary Range")

        wb.save(filename)
        return filename
    
    def _add_headers(self, ws, headers):
        """Add formatted headers to worksheet"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
    
    def _add_summary_sheet(self, wb, user_id, expenses):
        """Add summary sheet to workbook"""
        ws = wb.create_sheet("Summary")
        
        ws['A1'] = "Expense Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        
        # Calculate summary data
        summary_30 = self.db.get_summary(user_id, days=30)
        summary_7 = self.db.get_summary(user_id, days=7)
        
        total_30 = sum(amount for _, amount, _ in summary_30)
        total_7 = sum(amount for _, amount, _ in summary_7)
        
        # Summary stats
        ws['A3'] = "Last 7 Days:"
        ws['B3'] = total_7
        ws['B3'].number_format = f'"{CURRENCY}"#,##0.00'
        ws['B3'].font = Font(bold=True)
        
        ws['A4'] = "Last 30 Days:"
        ws['B4'] = total_30
        ws['B4'].number_format = f'"{CURRENCY}"#,##0.00'
        ws['B4'].font = Font(bold=True)
        
        if total_30 > 0:
            ws['A5'] = "Daily Average (30d):"
            ws['B5'] = total_30 / 30
            ws['B5'].number_format = f'"{CURRENCY}"#,##0.00'
        
        # Category breakdown
        ws['A7'] = "Category Breakdown (30 Days)"
        ws['A7'].font = Font(bold=True, size=11)
        
        ws['A8'] = "Category"
        ws['B8'] = "Amount"
        ws['C8'] = "Count"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}8'].font = Font(bold=True)
            ws[f'{col}8'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for row_idx, (category, amount, count) in enumerate(summary_30, start=9):
            ws[f'A{row_idx}'] = category
            ws[f'B{row_idx}'] = amount
            ws[f'C{row_idx}'] = count
            ws[f'B{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 18
    
    def _add_monthly_breakdown(self, wb, user_id, expenses):
        """Add monthly breakdown sheet"""
        ws = wb.create_sheet("Monthly Breakdown")
        
        ws['A1'] = "Monthly Breakdown"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        
        # Group by month
        monthly_data = {}
        for _, amount, category, _, date, _ in expenses:
            date_obj = self._parse_to_ist(date)
            month_key = date_obj.strftime("%Y-%m")
            if month_key not in monthly_data:
                monthly_data[month_key] = 0
            monthly_data[month_key] += amount
        
        ws['A3'] = "Month"
        ws['B3'] = "Total Spent"
        
        for col in ['A', 'B']:
            ws[f'{col}3'].font = Font(bold=True)
            ws[f'{col}3'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for row_idx, (month, total) in enumerate(sorted(monthly_data.items()), start=4):
            ws[f'A{row_idx}'] = month
            ws[f'B{row_idx}'] = total
            ws[f'B{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
        
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 24
    
    def _add_detailed_sheet(self, wb, expenses, sheet_name="Detailed", user_id=None):
        """Add detailed transactions sheet"""
        ws = wb.create_sheet(sheet_name)
        
        # Headers
        headers = [
            "ID",
            "Date",
            "Category",
            "Amount",
            "Description",
            "Source",
            "To",
            "From",
            "Amount",
            "Date/Time",
            "UPI transaction id",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.border = self.thin_border

        upi_meta_map = {}
        if user_id is not None:
            upi_meta_map = self._get_upi_meta_map(user_id, [exp_id for exp_id, *_ in expenses])

        # Data
        for seq, (exp_id, amount, category, description, date, source) in enumerate(expenses, start=1):
            row_idx = seq + 1
            date_obj = self._parse_to_ist(date)
            to_value, from_value, upi_amount, upi_date_time, upi_txn_id = self._extract_upi_excel_columns(
                upi_meta_map.get(exp_id, {}),
                amount,
                date_obj,
            )
            ws[f'A{row_idx}'] = seq
            ws[f'B{row_idx}'] = date_obj.strftime("%d-%m-%Y %H:%M")
            ws[f'C{row_idx}'] = category
            ws[f'D{row_idx}'] = amount
            ws[f'E{row_idx}'] = self._description_for_excel(description, category, source)
            ws[f'F{row_idx}'] = source if source else "text"
            ws[f'G{row_idx}'] = to_value
            ws[f'H{row_idx}'] = from_value
            ws[f'I{row_idx}'] = upi_amount
            ws[f'J{row_idx}'] = upi_date_time
            ws[f'K{row_idx}'] = upi_txn_id

            ws[f'D{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
            if upi_amount != "":
                ws[f'I{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'

            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                ws[f'{col}{row_idx}'].border = self.thin_border
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 46
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 50
        ws.column_dimensions['H'].width = 50
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 26
        ws.column_dimensions['K'].width = 50

    def _add_pattern_summary_sheet(self, wb, expenses, sheet_name="Pattern Summary"):
        """Add summary of extracted pattern names with amount/count."""
        ws = wb.create_sheet(sheet_name)

        pattern_totals = {}
        for _, amount, category, description, date, source in expenses:
            patterns = self._extract_pattern_list(description, category)
            if not patterns:
                patterns = ["other"]
            for pattern in patterns:
                if pattern not in pattern_totals:
                    pattern_totals[pattern] = {"amount": 0.0, "count": 0}
                pattern_totals[pattern]["amount"] += float(amount or 0)
                pattern_totals[pattern]["count"] += 1

        if not pattern_totals:
            ws['A1'] = "No pattern data found"
            return

        headers = ["Pattern Name", "Total Amount", "Count"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_idx, (pattern, data) in enumerate(
            sorted(pattern_totals.items(), key=lambda item: item[1]["amount"], reverse=True),
            start=2,
        ):
            ws[f'A{row_idx}'] = pattern
            ws[f'B{row_idx}'] = data["amount"]
            ws[f'C{row_idx}'] = data["count"]
            ws[f'B{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'

            for col in ['A', 'B', 'C']:
                ws[f'{col}{row_idx}'].border = self.thin_border

        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 16

    def _add_bill_totals_sheet(self, wb, user_id, days=None, start_date=None, end_date=None, sheet_name="Bill Totals"):
        """Add bill totals sheet with only subtotal/total/grand total pattern rows."""
        ws = wb.create_sheet(sheet_name)

        pattern_map = {
            "bill subtotal": "Subtotal",
            "subtotal": "Subtotal",
            "bill total": "Total",
            "total": "Total",
            "bill grand total": "Grand Total",
            "grand total": "Grand Total",
        }

        bill_rows = []
        for date, _, description, amount, source, _ in self._get_bill_analysis_rows(
            user_id,
            days=days,
            start_date=start_date,
            end_date=end_date,
        ):
            key = (description or "").strip().lower()
            pattern_name = pattern_map.get(key)
            if not pattern_name:
                continue
            bill_rows.append((date, pattern_name, amount, source if source else "image"))

        if not bill_rows:
            ws['A1'] = "No bill total pattern entries found"
            return

        headers = ["Date", "Pattern", "Amount", "Source"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_idx, (date, pattern_name, amount, source) in enumerate(bill_rows, start=2):
            date_obj = self._parse_to_ist(date)
            ws[f'A{row_idx}'] = date_obj.strftime("%d-%m-%Y %H:%M")
            ws[f'B{row_idx}'] = pattern_name
            ws[f'C{row_idx}'] = amount
            ws[f'D{row_idx}'] = source

            ws[f'C{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'

            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_idx}'].border = self.thin_border

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 26
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def _get_bill_analysis_rows(self, user_id, days=None, start_date=None, end_date=None):
        """Fetch bill-analysis entries from DB (bill subtotal/total labels)."""
        conn = sqlite3.connect(self.db.db_path)
        cursor = conn.cursor()

        labels = ("bill subtotal", "bill total", "bill grand total", "bill amount")

        if start_date and end_date:
            query = """
                SELECT date, category, description, amount, source, transaction_id
                FROM expenses
                WHERE user_id = ?
                  AND lower(COALESCE(description, '')) IN (?, ?, ?, ?)
                  AND date(date) BETWEEN date(?) AND date(?)
                ORDER BY date DESC, id DESC
            """
            cursor.execute(query, (user_id, *labels, start_date, end_date))
        elif days:
            query = """
                SELECT date, category, description, amount, source, transaction_id
                FROM expenses
                WHERE user_id = ?
                  AND lower(COALESCE(description, '')) IN (?, ?, ?, ?)
                  AND date >= datetime('now', '-' || ? || ' days')
                ORDER BY date DESC, id DESC
            """
            cursor.execute(query, (user_id, *labels, days))
        else:
            query = """
                SELECT date, category, description, amount, source, transaction_id
                FROM expenses
                WHERE user_id = ?
                  AND lower(COALESCE(description, '')) IN (?, ?, ?, ?)
                ORDER BY date DESC, id DESC
            """
            cursor.execute(query, (user_id, *labels))

        rows = cursor.fetchall()
        conn.close()
        return rows

    def _get_bill_item_rows(self, user_id, days=None, start_date=None, end_date=None):
        """Fetch item-level receipt rows (excluding bill subtotal/total meta labels)."""
        conn = sqlite3.connect(self.db.db_path)
        cursor = conn.cursor()

        labels = ("bill subtotal", "bill total", "bill grand total", "bill amount")
        if start_date and end_date:
            query = """
                SELECT date, category, description, amount, source, transaction_id
                FROM expenses
                WHERE user_id = ?
                  AND source = 'image'
                  AND lower(COALESCE(description, '')) NOT IN (?, ?, ?, ?)
                  AND date(date) BETWEEN date(?) AND date(?)
                ORDER BY date DESC, id DESC
            """
            cursor.execute(query, (user_id, *labels, start_date, end_date))
        elif days:
            query = """
                SELECT date, category, description, amount, source, transaction_id
                FROM expenses
                WHERE user_id = ?
                  AND source = 'image'
                  AND lower(COALESCE(description, '')) NOT IN (?, ?, ?, ?)
                  AND date >= datetime('now', '-' || ? || ' days')
                ORDER BY date DESC, id DESC
            """
            cursor.execute(query, (user_id, *labels, days))
        else:
            query = """
                SELECT date, category, description, amount, source, transaction_id
                FROM expenses
                WHERE user_id = ?
                  AND source = 'image'
                  AND lower(COALESCE(description, '')) NOT IN (?, ?, ?, ?)
                ORDER BY date DESC, id DESC
            """
            cursor.execute(query, (user_id, *labels))

        rows = cursor.fetchall()
        conn.close()
        return rows

    def _add_bill_items_sheet(self, wb, user_id, days=None, start_date=None, end_date=None, sheet_name="Bill Items"):
        """Add item-level receipt sheet with quantity, category, and amount."""
        ws = wb.create_sheet(sheet_name)
        rows = self._get_bill_item_rows(
            user_id,
            days=days,
            start_date=start_date,
            end_date=end_date,
        )

        if not rows:
            ws['A1'] = "No bill item entries found"
            return

        headers = ["Date", "Bill Ref", "Item Name", "Qty", "Category", "Amount", "Source"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_idx, (date, category, description, amount, source, bill_ref) in enumerate(rows, start=2):
            raw_desc = (description or "").strip()
            item_name = raw_desc.split("|")[0].strip() if raw_desc else "Receipt Item"
            qty_match = re.search(r"\bqty\s*:\s*([0-9]+(?:\.[0-9]+)?)", raw_desc, flags=re.IGNORECASE)
            qty = None
            if qty_match:
                qty_number = float(qty_match.group(1))
                qty = int(qty_number) if qty_number.is_integer() else qty_number

            date_obj = self._parse_to_ist(date)
            ws[f'A{row_idx}'] = date_obj.strftime("%d-%m-%Y %H:%M")
            ws[f'B{row_idx}'] = bill_ref or "-"
            ws[f'C{row_idx}'] = item_name
            ws[f'D{row_idx}'] = qty if qty is not None else ""
            ws[f'E{row_idx}'] = category or "Other"
            ws[f'F{row_idx}'] = amount
            ws[f'G{row_idx}'] = source or "image"

            ws[f'F{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'

            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws[f'{col}{row_idx}'].border = self.thin_border

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 34
        ws.column_dimensions['C'].width = 44
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 24
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 18

    def _add_bill_analysis_sheet(self, wb, user_id, days=None, start_date=None, end_date=None, sheet_name="Bill Analysis"):
        """Add bill analysis sheet with Category + Subtotal/Total/Grand Total/Amount."""
        ws = wb.create_sheet(sheet_name)
        rows = self._get_bill_analysis_rows(
            user_id,
            days=days,
            start_date=start_date,
            end_date=end_date,
        )

        if not rows:
            ws['A1'] = "No bill analysis entries found"
            return

        grouped = {}
        for date, category, description, amount, source, bill_ref in rows:
            key = bill_ref or f"{date}|{category}"
            if key not in grouped:
                grouped[key] = {
                    "date": date,
                    "category": category or "Other",
                    "subtotal": None,
                    "total": None,
                    "grand_total": None,
                    "amount": None,
                    "source": source or "image",
                    "bill_ref": bill_ref or "-",
                }

            desc = (description or "").strip().lower()
            if desc == "bill subtotal":
                grouped[key]["subtotal"] = amount
            elif desc == "bill total":
                grouped[key]["total"] = amount
            elif desc == "bill grand total":
                grouped[key]["grand_total"] = amount
            elif desc == "bill amount":
                grouped[key]["amount"] = amount

        headers = ["Date", "Category", "Subtotal", "Total", "Grand Total", "Amount", "Source", "Bill Ref"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_idx, data in enumerate(grouped.values(), start=2):
            date_obj = self._parse_to_ist(data["date"])
            ws[f'A{row_idx}'] = date_obj.strftime("%d-%m-%Y %H:%M")
            ws[f'B{row_idx}'] = data["category"]
            ws[f'C{row_idx}'] = data["subtotal"]
            ws[f'D{row_idx}'] = data["total"]
            ws[f'E{row_idx}'] = data["grand_total"]
            ws[f'F{row_idx}'] = data["amount"]
            ws[f'G{row_idx}'] = data["source"]
            ws[f'H{row_idx}'] = data["bill_ref"]

            for col in ['C', 'D', 'E', 'F']:
                ws[f'{col}{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'

            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws[f'{col}{row_idx}'].border = self.thin_border

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 34

    def _get_upi_rows(self, user_id, days=None, start_date=None, end_date=None):
        """Fetch UPI screenshot rows with extracted metadata."""
        conn = sqlite3.connect(self.db.db_path)
        cursor = conn.cursor()

        base_where = """
            user_id = ?
            AND (
                lower(COALESCE(source, '')) = 'online_payment'
                OR lower(COALESCE(payment_method, '')) = 'upi'
                OR COALESCE(upi_to, '') <> ''
                OR COALESCE(upi_from, '') <> ''
            )
        """

        if start_date and end_date:
            query = f"""
                SELECT date, amount, description, source, transaction_id, upi_to, upi_from, transaction_time
                FROM expenses
                WHERE {base_where}
                  AND date(date) BETWEEN date(?) AND date(?)
                ORDER BY date DESC, id DESC
            """
            cursor.execute(query, (user_id, start_date, end_date))
        elif days:
            query = f"""
                SELECT date, amount, description, source, transaction_id, upi_to, upi_from, transaction_time
                FROM expenses
                WHERE {base_where}
                  AND date >= datetime('now', '-' || ? || ' days')
                ORDER BY date DESC, id DESC
            """
            cursor.execute(query, (user_id, days))
        else:
            query = f"""
                SELECT date, amount, description, source, transaction_id, upi_to, upi_from, transaction_time
                FROM expenses
                WHERE {base_where}
                ORDER BY date DESC, id DESC
            """
            cursor.execute(query, (user_id,))

        rows = cursor.fetchall()
        conn.close()
        return rows

    def _add_upi_details_sheet(self, wb, user_id, days=None, start_date=None, end_date=None, sheet_name="UPI Details"):
        """Add UPI transaction detail sheet for extracted screenshot metadata."""
        ws = wb.create_sheet(sheet_name)
        rows = self._get_upi_rows(
            user_id,
            days=days,
            start_date=start_date,
            end_date=end_date,
        )

        headers = [
            "Logged Date",
            "Txn Date/Time",
            "Amount",
            "To",
            "From",
            "UPI Transaction ID",
            "Source",
            "Description",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if not rows:
            ws['A2'] = "No UPI screenshot entries found"
            ws.merge_cells("A2:H2")
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A2'].border = self.thin_border
            ws.column_dimensions['A'].width = 24
            ws.column_dimensions['B'].width = 26
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 28
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 40
            return

        for row_idx, (date, amount, description, source, txn_id, upi_to, upi_from, txn_time) in enumerate(rows, start=2):
            date_obj = self._parse_to_ist(date)
            txn_display = self._normalize_txn_datetime_display(txn_time, date_obj)
            ws[f'A{row_idx}'] = date_obj.strftime("%d-%m-%Y %H:%M")
            ws[f'B{row_idx}'] = txn_display
            ws[f'C{row_idx}'] = amount
            ws[f'D{row_idx}'] = upi_to or ""
            ws[f'E{row_idx}'] = upi_from or ""
            ws[f'F{row_idx}'] = txn_id or ""
            ws[f'G{row_idx}'] = source or ""
            ws[f'H{row_idx}'] = description or ""

            ws[f'C{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'

            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws[f'{col}{row_idx}'].border = self.thin_border

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 26
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 28
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 40
